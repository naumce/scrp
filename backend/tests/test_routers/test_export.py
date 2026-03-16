import csv
import io
import json

from openpyxl import load_workbook

from db.models import Business, EnrichmentResult


def _setup_project_with_businesses(client, session):
    resp = client.post("/api/projects/", json={
        "name": "Test", "keyword": "k", "location": "L"
    })
    pid = resp.json()["data"]["id"]

    biz1 = Business(project_id=pid, name="Biz1", website="https://biz1.com")
    biz2 = Business(project_id=pid, name="Biz2", website="https://biz2.com")
    session.add(biz1)
    session.add(biz2)
    session.commit()
    session.refresh(biz1)
    session.refresh(biz2)

    return pid, [biz1, biz2]


def test_export_csv(client, session):
    pid, businesses = _setup_project_with_businesses(client, session)

    resp = client.post("/api/export", json={
        "project_id": pid,
        "format": "csv",
    })
    assert resp.status_code == 200
    assert "text/csv" in resp.headers["content-type"]

    reader = csv.reader(io.StringIO(resp.text))
    rows = list(reader)
    assert len(rows) == 3  # header + 2 businesses
    assert rows[1][0] == "Biz1"
    assert rows[2][0] == "Biz2"


def test_export_excel(client, session):
    pid, businesses = _setup_project_with_businesses(client, session)

    resp = client.post("/api/export", json={
        "project_id": pid,
        "format": "excel",
    })
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 3
    assert rows[1][0] == "Biz1"


def test_export_with_enrichment(client, session):
    pid, businesses = _setup_project_with_businesses(client, session)

    er = EnrichmentResult(
        business_id=businesses[0].id,
        emails_json=json.dumps(["info@biz1.com"]),
        phones_json=json.dumps(["+15550100"]),
        social_links_json=json.dumps([]),
        status="success",
    )
    session.add(er)
    session.commit()

    resp = client.post("/api/export", json={
        "project_id": pid,
        "format": "csv",
    })
    rows = list(csv.reader(io.StringIO(resp.text)))
    assert "info@biz1.com" in rows[1][8]  # Emails column


def test_export_selected_ids(client, session):
    pid, businesses = _setup_project_with_businesses(client, session)

    resp = client.post("/api/export", json={
        "project_id": pid,
        "format": "csv",
        "business_ids": [businesses[0].id],
    })
    rows = list(csv.reader(io.StringIO(resp.text)))
    assert len(rows) == 2  # header + 1 selected


def test_export_no_businesses(client):
    resp = client.post("/api/projects/", json={
        "name": "Empty", "keyword": "k", "location": "L"
    })
    pid = resp.json()["data"]["id"]

    resp = client.post("/api/export", json={
        "project_id": pid,
        "format": "csv",
    })
    body = resp.json()
    assert body["success"] is False
    assert "No businesses" in body["error"]


def test_export_invalid_business_ids(client, session):
    pid, _ = _setup_project_with_businesses(client, session)

    resp = client.post("/api/export", json={
        "project_id": pid,
        "format": "csv",
        "business_ids": [99999],
    })
    body = resp.json()
    assert body["success"] is False
    assert "No matching" in body["error"]
