import csv
import io

from openpyxl import load_workbook

from db.models import Business, EnrichmentResult
from services.export_service import export_csv, export_excel, _build_row, COLUMNS


def _make_business(**overrides):
    defaults = dict(
        id=1,
        project_id=1,
        name="Test Biz",
        address="123 Main St",
        phone="+15551234",
        website="https://test.com",
        category="restaurant",
        rating=4.5,
        reviews=100,
        is_favorite=True,
        maps_url="https://maps.example.com",
    )
    defaults.update(overrides)
    return Business(**defaults)


def _make_enrichment(**overrides):
    defaults = dict(
        id=1,
        business_id=1,
        emails_json='["info@test.com"]',
        phones_json='["+15559999"]',
        social_links_json='["https://linkedin.com/test"]',
        contact_page="https://test.com/contact",
        status="success",
    )
    defaults.update(overrides)
    return EnrichmentResult(**defaults)


def test_build_row_with_enrichment():
    biz = _make_business()
    enrich = _make_enrichment()
    row = _build_row(biz, enrich)

    assert row.name == "Test Biz"
    assert row.address == "123 Main St"
    assert row.emails == "info@test.com"
    assert row.enriched_phones == "+15559999"
    assert row.social_links == "https://linkedin.com/test"
    assert row.contact_page == "https://test.com/contact"
    assert row.is_favorite == "Yes"
    assert row.rating == "4.5"


def test_build_row_without_enrichment():
    biz = _make_business(is_favorite=False, rating=None, reviews=None)
    row = _build_row(biz, None)

    assert row.emails == ""
    assert row.enriched_phones == ""
    assert row.social_links == ""
    assert row.contact_page == ""
    assert row.enrichment_status == ""
    assert row.is_favorite == "No"
    assert row.rating == ""
    assert row.reviews == ""


def test_export_csv_produces_valid_csv():
    biz = _make_business()
    enrich = _make_enrichment()
    data = export_csv([biz], {1: enrich})

    reader = csv.reader(io.StringIO(data.decode("utf-8")))
    rows = list(reader)

    # Header + 1 data row
    assert len(rows) == 2
    assert rows[0] == [col[0] for col in COLUMNS]
    assert rows[1][0] == "Test Biz"
    assert "info@test.com" in rows[1][8]


def test_export_csv_empty_list():
    data = export_csv([], {})
    reader = csv.reader(io.StringIO(data.decode("utf-8")))
    rows = list(reader)
    assert len(rows) == 1  # Header only


def test_export_csv_multiple_businesses():
    biz1 = _make_business(id=1, name="Biz One")
    biz2 = _make_business(id=2, name="Biz Two")
    data = export_csv([biz1, biz2], {})

    reader = csv.reader(io.StringIO(data.decode("utf-8")))
    rows = list(reader)
    assert len(rows) == 3
    assert rows[1][0] == "Biz One"
    assert rows[2][0] == "Biz Two"


def test_export_excel_produces_valid_workbook():
    biz = _make_business()
    enrich = _make_enrichment()
    data = export_excel([biz], {1: enrich})

    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    assert len(rows) == 2
    assert rows[0] == tuple(col[0] for col in COLUMNS)
    assert rows[1][0] == "Test Biz"
    assert "info@test.com" in rows[1][8]


def test_export_excel_header_bold():
    data = export_excel([_make_business()], {})
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    for cell in ws[1]:
        assert cell.font.bold is True


def test_export_excel_empty_list():
    data = export_excel([], {})
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 1  # Header only
