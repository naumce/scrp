"""
End-to-end integration test: create project → add businesses → enrich → export.
Tests the full data flow through routers without hitting external APIs.
"""
import csv
import io
import json
from unittest.mock import patch, AsyncMock

from db.models import Business, EnrichmentResult
from services.scraper_service import ScrapeResult


def test_full_flow(client, session):
    """
    E2E: create project → store businesses → enrich → check results → export CSV.
    """
    # 1. Create project
    resp = client.post("/api/projects/", json={
        "name": "E2E Test",
        "keyword": "restaurants",
        "location": "Chicago",
        "radius": 10,
        "max_results": 5,
    })
    assert resp.status_code == 200
    project = resp.json()["data"]
    pid = project["id"]
    assert project["name"] == "E2E Test"

    # 2. Verify project appears in list
    resp = client.get("/api/projects/")
    assert any(p["id"] == pid for p in resp.json()["data"])

    # 3. Add businesses directly (bypassing Overpass to avoid external calls)
    biz1 = Business(
        project_id=pid,
        name="Joe's Pizza",
        address="123 Main St, Chicago",
        phone="+13125551234",
        website="https://joespizza.com",
        category="restaurant",
        rating=4.5,
        reviews=200,
    )
    biz2 = Business(
        project_id=pid,
        name="Thai Palace",
        address="456 Oak Ave, Chicago",
        website="https://thaipalace.com",
        category="restaurant",
        rating=4.2,
        reviews=85,
    )
    biz3 = Business(
        project_id=pid,
        name="No Website Cafe",
        address="789 Elm St",
        category="restaurant",
    )
    session.add_all([biz1, biz2, biz3])
    session.commit()
    for b in [biz1, biz2, biz3]:
        session.refresh(b)

    # 4. Verify businesses show up in API
    resp = client.get(f"/api/businesses/?project_id={pid}")
    businesses = resp.json()["data"]
    assert len(businesses) == 3

    # 5. Toggle favorite
    resp = client.patch(f"/api/businesses/{biz1.id}/favorite")
    assert resp.json()["data"]["is_favorite"] is True

    # 6. Add enrichment results (simulating enrichment completion)
    er1 = EnrichmentResult(
        business_id=biz1.id,
        emails_json=json.dumps(["info@joespizza.com", "orders@joespizza.com"]),
        phones_json=json.dumps(["+13125551234"]),
        social_links_json=json.dumps(["https://facebook.com/joespizza"]),
        contact_page="https://joespizza.com/contact",
        status="success",
    )
    er2 = EnrichmentResult(
        business_id=biz2.id,
        emails_json=json.dumps(["hello@thaipalace.com"]),
        phones_json=json.dumps([]),
        social_links_json=json.dumps([
            "https://instagram.com/thaipalace",
            "https://facebook.com/thaipalace",
        ]),
        contact_page="https://thaipalace.com/about",
        status="success",
    )
    er3 = EnrichmentResult(
        business_id=biz3.id,
        status="failed",
        error_message="No website URL",
    )
    session.add_all([er1, er2, er3])
    session.commit()

    # 7. Verify enrichment results via API
    resp = client.get(f"/api/enrichment/results/{pid}")
    results = resp.json()["data"]
    assert len(results) == 3
    success_results = [r for r in results if r["status"] == "success"]
    assert len(success_results) == 2

    # 8. Export all as CSV
    resp = client.post("/api/export", json={
        "project_id": pid,
        "format": "csv",
    })
    assert resp.status_code == 200
    assert "text/csv" in resp.headers["content-type"]

    rows = list(csv.reader(io.StringIO(resp.text)))
    assert len(rows) == 4  # header + 3 businesses

    # Verify data integrity in export
    header = rows[0]
    name_idx = header.index("Name")
    email_idx = header.index("Emails")
    fav_idx = header.index("Favorite")
    status_idx = header.index("Enrichment Status")

    # Find Joe's Pizza row
    joes_row = next(r for r in rows[1:] if r[name_idx] == "Joe's Pizza")
    assert "info@joespizza.com" in joes_row[email_idx]
    assert "orders@joespizza.com" in joes_row[email_idx]
    assert joes_row[fav_idx] == "Yes"
    assert joes_row[status_idx] == "success"

    # Find No Website Cafe
    cafe_row = next(r for r in rows[1:] if r[name_idx] == "No Website Cafe")
    assert cafe_row[email_idx] == ""
    assert cafe_row[status_idx] == "failed"

    # 9. Export only selected businesses as Excel
    resp = client.post("/api/export", json={
        "project_id": pid,
        "format": "excel",
        "business_ids": [biz1.id, biz2.id],
    })
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    excel_rows = list(ws.iter_rows(values_only=True))
    assert len(excel_rows) == 3  # header + 2 selected businesses

    # 10. Verify settings round-trip
    resp = client.get("/api/settings")
    assert resp.json()["data"]["default_radius"] == "50"

    resp = client.put("/api/settings", json={"default_radius": "25"})
    assert resp.json()["data"]["default_radius"] == "25"

    resp = client.get("/api/settings")
    assert resp.json()["data"]["default_radius"] == "25"

    # 11. Delete project
    resp = client.delete(f"/api/projects/{pid}")
    assert resp.json()["success"] is True
